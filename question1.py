
from openpyxl import load_workbook

wb = load_workbook('sagatave_eksamenam (1).xlsx')
ws = wb['Lapa_0']

count = 0
for row in ws.iter_rows(min_row=3, values_only=True):
    adrese = row[3]
    skaits = row[11]
    if adrese and isinstance(skaits, (int, float)) and adrese.startswith('Ain') and skaits < 40:
        count += 1

print("Answer 1:", count)
