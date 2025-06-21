
from openpyxl import load_workbook
from datetime import datetime

wb = load_workbook('sagatave_eksamenam (1).xlsx')
ws = wb['Lapa_0']

count = 0
for row in ws.iter_rows(min_row=3, values_only=True):
    prioritize = row[7]
    date = row[9]
    if prioritize == 'High' and isinstance(date, datetime) and date.year == 2015:
        count += 1

print("Answer 2:", count)
