
from openpyxl import load_workbook
import math

wb = load_workbook('sagatave_eksamenam (1).xlsx')
ws = wb['Lapa_0']

total = 0
count = 0
for row in ws.iter_rows(min_row=3, values_only=True):
    produkts = row[8]
    cena = row[10]
    if produkts and 'LaserJet' in produkts and isinstance(cena, (int, float)):
        total += cena
        count += 1

average = math.floor(total / count) if count else 0
print("Answer 4:", average)
