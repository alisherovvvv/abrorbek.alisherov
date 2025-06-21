
from openpyxl import load_workbook

wb = load_workbook('sagatave_eksamenam (1).xlsx')
ws = wb['Lapa_0']

count = 0
for row in ws.iter_rows(min_row=3, values_only=True):
    adrese = row[3]
    pilseta = row[4]
    if adrese == 'Adulienas iela' and pilseta in ['Valmiera', 'Saulkrasti']:
        count += 1

print("Answer 3:", count)
