
from openpyxl import load_workbook
import math

wb = load_workbook('sagatave_eksamenam (1).xlsx')
ws = wb['Lapa_0']

total = 0
for row in ws.iter_rows(min_row=3, values_only=True):
    klients = row[5]
    skaits = row[11]
    cena = row[10]
    pieg_cena = row[12]
    if klients == 'Korporatīvais' and isinstance(skaits, (int, float)) and 40 <= skaits <= 50:
        if all(isinstance(x, (int, float)) for x in [cena, skaits, pieg_cena]):
            total += cena * skaits + pieg_cena

print("Answer 5:", math.floor(total))
